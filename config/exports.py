import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse
from datetime import datetime

VERDE = "10B981"
OSCURO = "1a1a2e"
GRIS = "f8fafc"

def estilo_header(ws, fila, columnas):
    for col, titulo in enumerate(columnas, 1):
        cell = ws.cell(row=fila, column=col, value=titulo)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=OSCURO)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            bottom=Side(style="thin", color=VERDE)
        )

def estilo_fila(ws, fila, par=False):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=fila, column=col)
        cell.fill = PatternFill("solid", fgColor="F1F5F9" if par else "FFFFFF")
        cell.alignment = Alignment(vertical="center")

def autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

def exportar_clientes():
    from clientes.models import Cliente
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"
    ws.row_dimensions[1].height = 28

    cols = ["ID", "Nombre", "Documento", "Teléfono", "WhatsApp", "Correo",
            "Ciudad", "Tipo", "Puntos", "Activo", "Fecha Registro"]
    estilo_header(ws, 1, cols)

    for i, c in enumerate(Cliente.objects.all().order_by('-fecha_registro'), 2):
        ws.append([
            c.id, c.nombre, c.documento, c.telefono, c.whatsapp,
            c.correo, c.ciudad, c.tipo, c.puntos,
            "Sí" if c.activo else "No",
            c.fecha_registro.strftime("%Y-%m-%d %H:%M") if c.fecha_registro else ""
        ])
        estilo_fila(ws, i, i % 2 == 0)

    autofit(ws)
    return wb

def exportar_ordenes():
    from servicios.models import OrdenTrabajo
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ordenes"
    ws.row_dimensions[1].height = 28

    cols = ["Código", "Cliente", "Documento", "Teléfono", "Placa", "Marca",
            "Estado", "Prioridad", "Técnico", "Costo Final", "Fecha Ingreso", "Fecha Salida"]
    estilo_header(ws, 1, cols)

    for i, o in enumerate(OrdenTrabajo.objects.select_related('cliente','vehiculo').all().order_by('-fecha_ingreso'), 2):
        ws.append([
            o.codigo, o.cliente.nombre, o.cliente.documento, o.cliente.telefono,
            o.vehiculo.placa, f"{o.vehiculo.marca} {o.vehiculo.linea}",
            o.estado, o.prioridad, o.tecnico, float(o.costo_final),
            o.fecha_ingreso.strftime("%Y-%m-%d %H:%M") if o.fecha_ingreso else "",
            o.fecha_salida.strftime("%Y-%m-%d %H:%M") if o.fecha_salida else ""
        ])
        estilo_fila(ws, i, i % 2 == 0)

    autofit(ws)
    return wb

def exportar_inventario():
    from inventario.models import Producto
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"
    ws.row_dimensions[1].height = 28

    cols = ["SKU", "Nombre", "Categoría", "Proveedor", "Costo", "Precio Venta",
            "Margen %", "Stock Actual", "Stock Mínimo", "Stock Crítico", "Estado"]
    estilo_header(ws, 1, cols)

    for i, p in enumerate(Producto.objects.select_related('categoria','proveedor').all(), 2):
        ws.append([
            p.sku, p.nombre,
            p.categoria.nombre if p.categoria else "",
            p.proveedor.nombre if p.proveedor else "",
            float(p.costo), float(p.precio_venta), float(p.margen),
            p.stock_actual, p.stock_minimo, p.stock_critico, p.estado_stock
        ])
        estilo_fila(ws, i, i % 2 == 0)

    autofit(ws)
    return wb

def exportar_facturas():
    from contabilidad.models import Factura
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Facturas"
    ws.row_dimensions[1].height = 28

    cols = ["Número", "Cliente", "Documento", "Estado", "Método Pago",
            "Subtotal", "Descuento", "Total", "Fecha Emisión", "Fecha Pago"]
    estilo_header(ws, 1, cols)

    for i, f in enumerate(Factura.objects.select_related('cliente').all().order_by('-fecha_emision'), 2):
        ws.append([
            f.numero, f.cliente.nombre, f.cliente.documento,
            f.estado, f.metodo_pago,
            float(f.subtotal), float(f.descuento), float(f.total),
            f.fecha_emision.strftime("%Y-%m-%d %H:%M") if f.fecha_emision else "",
            f.fecha_pago.strftime("%Y-%m-%d %H:%M") if f.fecha_pago else ""
        ])
        estilo_fila(ws, i, i % 2 == 0)

    autofit(ws)
    return wb

def exportar_gastos():
    from contabilidad.models import Gasto
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gastos"
    ws.row_dimensions[1].height = 28

    cols = ["ID", "Descripción", "Categoría", "Monto", "Fecha", "Comprobante"]
    estilo_header(ws, 1, cols)

    for i, g in enumerate(Gasto.objects.all().order_by('-fecha'), 2):
        ws.append([
            g.id, g.descripcion, g.categoria, float(g.monto),
            g.fecha.strftime("%Y-%m-%d") if g.fecha else "",
            g.comprobante
        ])
        estilo_fila(ws, i, i % 2 == 0)

    autofit(ws)
    return wb

def exportar_completo():
    """Exportar todo en un solo archivo con múltiples hojas"""
    from clientes.models import Cliente
    from servicios.models import OrdenTrabajo
    from inventario.models import Producto
    from contabilidad.models import Factura, Gasto

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Hoja 1: Clientes
    ws1 = wb.create_sheet("Clientes")
    cols1 = ["ID", "Nombre", "Documento", "Teléfono", "Correo", "Ciudad", "Tipo", "Puntos"]
    estilo_header(ws1, 1, cols1)
    for i, c in enumerate(Cliente.objects.all(), 2):
        ws1.append([c.id, c.nombre, c.documento, c.telefono, c.correo, c.ciudad, c.tipo, c.puntos])
        estilo_fila(ws1, i, i % 2 == 0)
    autofit(ws1)

    # Hoja 2: Órdenes
    ws2 = wb.create_sheet("Ordenes")
    cols2 = ["Código", "Cliente", "Placa", "Estado", "Prioridad", "Costo", "Fecha"]
    estilo_header(ws2, 1, cols2)
    for i, o in enumerate(OrdenTrabajo.objects.select_related('cliente','vehiculo').all(), 2):
        ws2.append([o.codigo, o.cliente.nombre, o.vehiculo.placa, o.estado,
                    o.prioridad, float(o.costo_final),
                    o.fecha_ingreso.strftime("%Y-%m-%d") if o.fecha_ingreso else ""])
        estilo_fila(ws2, i, i % 2 == 0)
    autofit(ws2)

    # Hoja 3: Inventario
    ws3 = wb.create_sheet("Inventario")
    cols3 = ["SKU", "Nombre", "Stock", "Costo", "Precio", "Margen %"]
    estilo_header(ws3, 1, cols3)
    for i, p in enumerate(Producto.objects.all(), 2):
        ws3.append([p.sku, p.nombre, p.stock_actual, float(p.costo), float(p.precio_venta), float(p.margen)])
        estilo_fila(ws3, i, i % 2 == 0)
    autofit(ws3)

    # Hoja 4: Facturas
    ws4 = wb.create_sheet("Facturas")
    cols4 = ["Número", "Cliente", "Estado", "Total", "Fecha"]
    estilo_header(ws4, 1, cols4)
    for i, f in enumerate(Factura.objects.select_related('cliente').all(), 2):
        ws4.append([f.numero, f.cliente.nombre, f.estado, float(f.total),
                    f.fecha_emision.strftime("%Y-%m-%d") if f.fecha_emision else ""])
        estilo_fila(ws4, i, i % 2 == 0)
    autofit(ws4)

    # Hoja 5: Gastos
    ws5 = wb.create_sheet("Gastos")
    cols5 = ["Descripción", "Categoría", "Monto", "Fecha"]
    estilo_header(ws5, 1, cols5)
    for i, g in enumerate(Gasto.objects.all(), 2):
        ws5.append([g.descripcion, g.categoria, float(g.monto),
                    g.fecha.strftime("%Y-%m-%d") if g.fecha else ""])
        estilo_fila(ws5, i, i % 2 == 0)
    autofit(ws5)

    return wb

def wb_to_response(wb, filename):
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
